from __future__ import annotations

from datetime import datetime
from email.utils import parsedate_to_datetime
from io import BytesIO

import feedparser
from openpyxl import load_workbook

from ..models import Company, RawRelease
from ..textutil import normalize_whitespace, parse_date
from .base import Extractor


class RssExtractor(Extractor):
    source_type = "rss"

    def fetch(self, company: Company, limit: int) -> list[RawRelease]:
        feed_url = company.config.get("feed_url") or company.list_url
        parsed = feedparser.parse(self.http.get_text(feed_url))
        results: list[RawRelease] = []
        for entry in parsed.entries[:limit]:
            title = normalize_whitespace(getattr(entry, "title", ""))
            link = getattr(entry, "link", "")
            if not title or not link:
                continue
            published = None
            if getattr(entry, "published", None):
                try:
                    published = parsedate_to_datetime(entry.published).date()
                except (TypeError, ValueError, IndexError):
                    published = parse_date(entry.published)
            summary = normalize_whitespace(
                getattr(entry, "summary", "") or getattr(entry, "description", "")
            )
            results.append(
                RawRelease(
                    company=company.name,
                    title=title,
                    url=link,
                    published_on=published,
                    summary=summary,
                    source_type=self.source_type,
                )
            )
        return results


class XlsxExtractor(Extractor):
    source_type = "xlsx"

    def fetch(self, company: Company, limit: int) -> list[RawRelease]:
        cfg = company.config
        xlsx_url = cfg.get("xlsx_url")
        if not xlsx_url:
            raise ValueError(f"{company.name}: xlsx_url is required")
        content = self._download_xlsx(xlsx_url, referer=company.list_url)
        wb = load_workbook(BytesIO(content), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        title_col = cfg.get("title_column", "title")
        date_col = cfg.get("date_column", "date")
        url_col = cfg.get("url_column", "url")
        try:
            title_idx = headers.index(title_col)
            date_idx = headers.index(date_col)
            url_idx = headers.index(url_col)
        except ValueError as exc:
            raise ValueError(f"{company.name}: unexpected xlsx headers {headers}") from exc

        results: list[RawRelease] = []
        for row in rows[1:]:
            if not row or len(row) <= max(title_idx, date_idx, url_idx):
                continue
            title = normalize_whitespace(str(row[title_idx] or ""))
            url = str(row[url_idx] or "").strip()
            if not title or not url:
                continue
            published = parse_date(row[date_idx])
            results.append(
                RawRelease(
                    company=company.name,
                    title=title,
                    url=url,
                    published_on=published,
                    summary="",
                    source_type=self.source_type,
                )
            )
            if len(results) >= limit:
                break
        return results

    def _download_xlsx(self, xlsx_url: str, referer: str) -> bytes:
        try:
            return self.http.get_bytes(xlsx_url, referer=referer)
        except Exception:
            # Some hosts block datacenter IPs/User-Agents on direct downloads.
            from playwright.sync_api import sync_playwright

            with sync_playwright() as p:
                browser = p.chromium.launch(headless=True)
                context = browser.new_context(
                    user_agent=self.http._client.headers["User-Agent"],
                    extra_http_headers={"Referer": referer},
                )
                page = context.new_page()
                page.goto(referer, wait_until="domcontentloaded", timeout=60000)
                response = page.request.get(xlsx_url)
                if response.status >= 400:
                    browser.close()
                    raise RuntimeError(
                        f"Failed to download xlsx via playwright: {response.status}"
                    )
                content = response.body()
                browser.close()
                return content

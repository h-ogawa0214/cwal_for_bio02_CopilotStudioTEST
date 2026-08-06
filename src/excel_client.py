from __future__ import annotations

import json
import time
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .dedupe import canonicalize_url, release_fingerprint
from .metrics import RunMetrics
from .models import Company, CuratedRelease, DecisionRecord
from .settings import Settings


COMPANY_HEADERS = [
    "company_name",
    "stock_code",
    "list_url",
    "enabled",
    "source_type",
    "crawl_mode",
    "config_json",
    "notes",
]

RELEASE_HEADERS = [
    "published_on",
    "company_name",
    "title",
    "paragraph",
    "url",
    "fetched_at",
    "decision_reason",
    "original_title",
    "reference_url",
]

DECISION_HEADERS = [
    "decided_at",
    "company_name",
    "published_on",
    "title",
    "url",
    "canonical_url",
    "fingerprint",
    "content_hash",
    "decision",
    "reason",
    "source_type",
    "model",
    "criteria_version",
]

METRICS_HEADERS = [
    "run_at",
    "candidates_seen",
    "candidates_new",
    "cache_hits",
    "kept",
    "discarded",
    "hard_discards",
    "heuristic_discards",
    "pending_review",
    "duplicates_skipped",
    "fetch_errors",
    "site_only",
    "tdnet_only",
    "matched_site_tdnet",
    "source_stats_json",
]

_SAVE_RETRIES = 3
_SAVE_RETRY_WAIT_SECONDS = 2.0


class ExcelWriteError(RuntimeError):
    """Raised when the workbook can't be read/written (e.g. open in Excel)."""


def _cell_to_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def _records(ws: Worksheet) -> list[dict[str, str]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    records: list[dict[str, str]] = []
    for row in rows[1:]:
        if row is None or all(v is None for v in row):
            continue
        record: dict[str, str] = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            value = row[idx] if idx < len(row) else None
            record[header] = _cell_to_str(value)
        records.append(record)
    return records


class ExcelClient:
    """OneDrive上の単一 .xlsx を companies/releases/decisions/metrics の
    4シート構成で読み書きする。SheetsClient と同じ公開インターフェースを維持する。
    """

    def __init__(self, settings: Settings) -> None:
        self.path = Path(settings.excel_file_path)
        self.companies_sheet_name = settings.companies_sheet
        self.releases_sheet_name = settings.releases_sheet
        self.decisions_sheet_name = settings.decisions_sheet
        self.metrics_sheet_name = settings.metrics_sheet

    # -- workbook I/O -----------------------------------------------------

    def _load_workbook(self) -> Workbook:
        if self.path.exists():
            try:
                return load_workbook(self.path)
            except PermissionError as exc:
                raise ExcelWriteError(
                    f"{self.path} を読み込めません。Excelで開いている場合は閉じてから再実行してください。"
                ) from exc
        self.path.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        wb.remove(wb.active)  # drop the default empty "Sheet"
        return wb

    def _save_workbook(self, wb: Workbook) -> None:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        tmp_path = self.path.with_name(self.path.stem + ".tmp" + self.path.suffix)
        last_error: Exception | None = None
        for attempt in range(1, _SAVE_RETRIES + 1):
            try:
                wb.save(tmp_path)
                tmp_path.replace(self.path)
                return
            except PermissionError as exc:
                last_error = exc
                if attempt < _SAVE_RETRIES:
                    time.sleep(_SAVE_RETRY_WAIT_SECONDS)
        raise ExcelWriteError(
            f"{self.path} に保存できません。Excelで開いている場合は閉じてから再実行してください。"
        ) from last_error

    def _ensure_sheet(self, wb: Workbook, title: str, headers: list[str]) -> Worksheet:
        if title in wb.sheetnames:
            ws = wb[title]
            existing = [c.value for c in ws[1]] if ws.max_row >= 1 else []
            existing_clean = [v for v in existing if v]
            if not existing_clean:
                for col, header in enumerate(headers, start=1):
                    ws.cell(row=1, column=col, value=header)
            else:
                missing = [h for h in headers if h not in existing_clean]
                if missing:
                    start_col = len(existing_clean) + 1
                    for offset, header in enumerate(missing):
                        ws.cell(row=1, column=start_col + offset, value=header)
            return ws
        ws = wb.create_sheet(title=title)
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)
        return ws

    def ensure_schema(self) -> None:
        wb = self._load_workbook()
        self._ensure_sheet(wb, self.companies_sheet_name, COMPANY_HEADERS)
        self._ensure_sheet(wb, self.releases_sheet_name, RELEASE_HEADERS)
        self._ensure_sheet(wb, self.decisions_sheet_name, DECISION_HEADERS)
        self._ensure_sheet(wb, self.metrics_sheet_name, METRICS_HEADERS)
        self._save_workbook(wb)

    # -- companies ----------------------------------------------------------

    def load_companies(self) -> list[Company]:
        wb = self._load_workbook()
        if self.companies_sheet_name not in wb.sheetnames:
            return []
        companies: list[Company] = []
        for row in _records(wb[self.companies_sheet_name]):
            name = str(row.get("company_name") or "").strip()
            list_url = str(row.get("list_url") or "").strip()
            source_type = str(row.get("source_type") or "").strip()
            if not (name and list_url and source_type):
                continue
            enabled_raw = str(row.get("enabled") or "TRUE").strip().lower()
            enabled = enabled_raw in {"1", "true", "yes", "y", "有効"}
            config_raw = str(row.get("config_json") or "").strip()
            config = json.loads(config_raw) if config_raw else {}
            crawl_mode_raw = str(row.get("crawl_mode") or "live").strip().lower()
            crawl_mode = "shadow" if crawl_mode_raw == "shadow" else "live"
            companies.append(
                Company(
                    name=name,
                    list_url=list_url,
                    enabled=enabled,
                    source_type=source_type,
                    stock_code=str(row.get("stock_code") or "").strip(),
                    crawl_mode=crawl_mode,
                    config=config,
                    notes=str(row.get("notes") or ""),
                )
            )
        return companies

    def seed_companies_if_empty(self, companies: Iterable[Company]) -> int:
        existing = self.load_companies()
        if existing:
            return 0
        wb = self._load_workbook()
        ws = self._ensure_sheet(wb, self.companies_sheet_name, COMPANY_HEADERS)
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        count = 0
        for company in companies:
            ws.append(self._company_row(company))
            count += 1
        self._save_workbook(wb)
        return count

    def sync_companies(self, companies: Iterable[Company]) -> dict[str, int]:
        """Sync YAML authority fields and append missing companies.

        YAML owns: list_url, source_type, config_json, stock_code, crawl_mode default.
        Excel may override: enabled, crawl_mode (if non-empty), notes.
        """
        self.ensure_schema()
        wb = self._load_workbook()
        ws = wb[self.companies_sheet_name]
        headers = [c.value for c in ws[1]]
        header_index = {name: idx for idx, name in enumerate(headers) if name}
        name_idx = header_index.get("company_name", 0)

        existing_names: set[str] = set()
        for row_number in range(2, ws.max_row + 1):
            value = ws.cell(row=row_number, column=name_idx + 1).value
            if value:
                existing_names.add(str(value).strip())

        appended = 0
        updated_fields = 0
        authority_fields = {
            "list_url": lambda c: c.list_url,
            "source_type": lambda c: c.source_type,
            "stock_code": lambda c: c.stock_code,
            "config_json": lambda c: json.dumps(c.config, ensure_ascii=False),
            "crawl_mode": lambda c: c.crawl_mode,
        }
        companies_by_name = {c.name: c for c in companies}

        for company in companies_by_name.values():
            if company.name not in existing_names:
                self._append_company_row(ws, header_index, company)
                existing_names.add(company.name)
                appended += 1

        for row_number in range(2, ws.max_row + 1):
            raw_name = ws.cell(row=row_number, column=name_idx + 1).value
            name = str(raw_name).strip() if raw_name else ""
            company = companies_by_name.get(name)
            if company is None:
                continue
            for field, getter in authority_fields.items():
                col_idx = header_index.get(field)
                if col_idx is None:
                    continue
                desired = getter(company)
                if field == "stock_code" and not desired:
                    continue
                cell = ws.cell(row=row_number, column=col_idx + 1)
                current = str(cell.value).strip() if cell.value else ""
                # Keep an explicit sheet crawl_mode once set.
                if field == "crawl_mode" and current:
                    continue
                if current == desired:
                    continue
                cell.value = desired
                updated_fields += 1

        self._save_workbook(wb)
        return {"appended": appended, "updated_fields": updated_fields}

    def _append_company_row(
        self, ws: Worksheet, header_index: dict[str, int], company: Company
    ) -> None:
        width = max(ws.max_column, len(header_index), len(COMPANY_HEADERS))
        row_values: list[object] = [None] * width
        for header, value in zip(COMPANY_HEADERS, self._company_row(company)):
            idx = header_index.get(header)
            if idx is not None:
                row_values[idx] = value
        ws.append(row_values)

    @staticmethod
    def _company_row(company: Company) -> list[str]:
        return [
            company.name,
            company.stock_code,
            company.list_url,
            "TRUE" if company.enabled else "FALSE",
            company.source_type,
            company.crawl_mode,
            json.dumps(company.config, ensure_ascii=False),
            company.notes,
        ]

    # -- releases -------------------------------------------------------------

    def existing_urls(self) -> set[str]:
        urls, _ = self.existing_release_keys()
        return urls

    def existing_release_keys(self) -> tuple[set[str], set[str]]:
        """Return known URLs and soft fingerprints already stored in releases."""
        wb = self._load_workbook()
        if self.releases_sheet_name not in wb.sheetnames:
            return set(), set()
        urls: set[str] = set()
        fingerprints: set[str] = set()
        for row in _records(wb[self.releases_sheet_name]):
            url = str(row.get("url") or "").strip()
            if url:
                urls.add(url)
                canonical = canonicalize_url(url)
                if canonical:
                    urls.add(canonical)
            company = str(row.get("company_name") or "").strip()
            published_on = str(row.get("published_on") or "").strip()
            title = str(row.get("original_title") or row.get("title") or "").strip()
            if company and title:
                fingerprints.add(release_fingerprint(company, published_on, title))
                display_title = str(row.get("title") or "").strip()
                if display_title and display_title != title:
                    fingerprints.add(
                        release_fingerprint(company, published_on, display_title)
                    )
        return urls, fingerprints

    def upsert_releases(self, releases: list[CuratedRelease]) -> int:
        if not releases:
            self.sort_releases_by_date()
            return 0
        wb = self._load_workbook()
        ws = self._ensure_sheet(wb, self.releases_sheet_name, RELEASE_HEADERS)
        headers = [c.value for c in ws[1]] or RELEASE_HEADERS
        url_index = headers.index("url")
        existing_rows: dict[str, int] = {}
        for row_number in range(2, ws.max_row + 1):
            value = ws.cell(row=row_number, column=url_index + 1).value
            if value:
                existing_rows[str(value).strip()] = row_number

        _, existing_fingerprints = self.existing_release_keys()
        rows_to_append: list[list[str]] = []
        written = 0
        for item in releases:
            fingerprint = release_fingerprint(
                item.company,
                item.published_on,
                item.original_title or item.title,
            )
            row = [
                item.published_on.isoformat(),
                item.company,
                item.title,
                item.paragraph,
                item.url,
                item.fetched_at.replace(tzinfo=timezone.utc).isoformat(),
                item.reason,
                item.original_title,
                item.reference_url,
            ]
            existing_row = existing_rows.get(item.url)
            if existing_row:
                for col_idx, value in enumerate(row, start=1):
                    ws.cell(row=existing_row, column=col_idx, value=value)
                existing_fingerprints.add(fingerprint)
                written += 1
            elif fingerprint in existing_fingerprints:
                continue
            else:
                rows_to_append.append(row)
                existing_fingerprints.add(fingerprint)
                written += 1

        for row in rows_to_append:
            ws.append(row)

        self._save_workbook(wb)
        self.sort_releases_by_date()
        return written

    def sort_releases_by_date(self) -> None:
        """Keep the releases sheet newest-first while leaving the header in place."""
        wb = self._load_workbook()
        if self.releases_sheet_name not in wb.sheetnames:
            return
        ws = wb[self.releases_sheet_name]
        if ws.max_row <= 2:
            return
        headers = [c.value for c in ws[1]]
        if "published_on" not in headers:
            return
        date_idx = headers.index("published_on")
        rows = [
            list(row)
            for row in ws.iter_rows(min_row=2, values_only=True)
            if row is not None and any(v is not None for v in row)
        ]
        rows.sort(key=lambda row: str(row[date_idx] or ""), reverse=True)
        for row_number, row in enumerate(rows, start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_number, column=col_idx, value=value)
        self._save_workbook(wb)

    def append_releases(self, releases: list[CuratedRelease]) -> int:
        return self.upsert_releases(releases)

    # -- decisions / metrics --------------------------------------------------

    def load_decision_cache(self, *, today: date | None = None) -> dict[str, DecisionRecord]:
        """Map fingerprint/canonical_url -> latest decision for same-day discard cache."""
        wb = self._load_workbook()
        if self.decisions_sheet_name not in wb.sheetnames:
            return {}
        cache: dict[str, DecisionRecord] = {}
        day = today or date.today()
        for row in _records(wb[self.decisions_sheet_name]):
            decision = str(row.get("decision") or "").strip().lower()
            if decision not in {"keep", "discard", "hard_discard"}:
                continue
            decided_raw = str(row.get("decided_at") or "").strip()
            try:
                decided_at = datetime.fromisoformat(decided_raw.replace("Z", "+00:00"))
            except ValueError:
                continue
            decided_day = decided_at.astimezone(timezone.utc).date()
            # Same-day discard cache prevents re-LLM across repeated same-day runs.
            if decision in {"discard", "hard_discard"} and decided_day != day:
                continue
            record = DecisionRecord(
                decided_at=decided_at,
                company=str(row.get("company_name") or "").strip(),
                published_on=str(row.get("published_on") or "").strip(),
                title=str(row.get("title") or "").strip(),
                url=str(row.get("url") or "").strip(),
                canonical_url=str(row.get("canonical_url") or "").strip(),
                fingerprint=str(row.get("fingerprint") or "").strip(),
                content_hash=str(row.get("content_hash") or "").strip(),
                decision=decision,
                reason=str(row.get("reason") or "").strip(),
                source_type=str(row.get("source_type") or "").strip(),
                model=str(row.get("model") or "").strip(),
                criteria_version=str(row.get("criteria_version") or "").strip(),
            )
            for key in (
                record.fingerprint,
                record.canonical_url,
                record.url,
                record.content_hash,
            ):
                if key:
                    cache[key] = record
        return cache

    def append_decisions(self, decisions: list[DecisionRecord]) -> int:
        if not decisions:
            return 0
        wb = self._load_workbook()
        ws = self._ensure_sheet(wb, self.decisions_sheet_name, DECISION_HEADERS)
        for item in decisions:
            ws.append(
                [
                    item.decided_at.replace(tzinfo=timezone.utc).isoformat(),
                    item.company,
                    item.published_on,
                    item.title,
                    item.url,
                    item.canonical_url,
                    item.fingerprint,
                    item.content_hash,
                    item.decision,
                    item.reason,
                    item.source_type,
                    item.model,
                    item.criteria_version,
                ]
            )
        self._save_workbook(wb)
        return len(decisions)

    def append_run_metrics(self, metrics: RunMetrics) -> None:
        wb = self._load_workbook()
        ws = self._ensure_sheet(wb, self.metrics_sheet_name, METRICS_HEADERS)
        ws.append(
            [
                metrics.started_at.replace(tzinfo=timezone.utc).isoformat(),
                metrics.candidates_seen,
                metrics.candidates_new,
                metrics.cache_hits,
                metrics.kept,
                metrics.discarded,
                metrics.hard_discards,
                metrics.heuristic_discards,
                metrics.pending_review,
                metrics.duplicates_skipped,
                metrics.fetch_errors,
                metrics.site_only,
                metrics.tdnet_only,
                metrics.matched_site_tdnet,
                json.dumps(metrics.source_stats, ensure_ascii=False),
            ]
        )
        self._save_workbook(wb)

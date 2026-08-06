from __future__ import annotations

import tempfile
import unittest
from datetime import date
from pathlib import Path

from openpyxl import Workbook

from src.dedupe import release_fingerprint
from src.excel_client import (
    COMPANY_HEADERS,
    RELEASE_HEADERS,
    ExcelClient,
)
from src.models import Company


def _client(path: Path) -> ExcelClient:
    client = ExcelClient.__new__(ExcelClient)
    client.path = path
    client.companies_sheet_name = "companies"
    client.releases_sheet_name = "releases"
    client.decisions_sheet_name = "decisions"
    client.metrics_sheet_name = "metrics"
    return client


class ExcelClientTests(unittest.TestCase):
    def setUp(self) -> None:
        self._tmpdir = tempfile.TemporaryDirectory()
        self.addCleanup(self._tmpdir.cleanup)
        self.path = Path(self._tmpdir.name) / "releases.xlsx"

    def test_company_headers_include_crawl_mode(self) -> None:
        self.assertIn("crawl_mode", COMPANY_HEADERS)

    def test_ensure_schema_creates_all_sheets(self) -> None:
        client = _client(self.path)
        client.ensure_schema()

        from openpyxl import load_workbook

        wb = load_workbook(self.path)
        self.assertEqual(
            set(wb.sheetnames),
            {"companies", "releases", "decisions", "metrics"},
        )
        self.assertEqual([c.value for c in wb["companies"][1]], COMPANY_HEADERS)

    def test_existing_release_keys_include_fingerprints(self) -> None:
        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("releases")
        ws.append(RELEASE_HEADERS)
        ws.append(
            [
                "2026-07-17",
                "レナサイエンス",
                "短い題",
                "リード文",
                "https://example.com/a",
                "2026-07-17T00:00:00+00:00",
                "",
                "局所進行非小細胞肺がんに対する第二相治験開始のお知らせ",
                "",
            ]
        )
        wb.save(self.path)

        client = _client(self.path)
        urls, fingerprints = client.existing_release_keys()

        self.assertIn("https://example.com/a", urls)
        self.assertIn(
            release_fingerprint(
                "レナサイエンス",
                "2026-07-17",
                "局所進行非小細胞肺がんに対する第二相治験開始のお知らせ",
            ),
            fingerprints,
        )

    def test_release_fingerprint_ignores_whitespace(self) -> None:
        left = release_fingerprint("協和キリン", date(2026, 6, 12), "新規診断 AML")
        right = release_fingerprint("協和キリン", "2026-06-12", "新規診断AML")
        self.assertEqual(left, right)

    def test_sync_companies_updates_authority_fields_and_preserves_new_mode(
        self,
    ) -> None:
        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("companies")
        ws.append(COMPANY_HEADERS)
        ws.append(["エーザイ", "4523", "old_url", "TRUE", "tdnet_only", "", "{}", ""])
        wb.save(self.path)

        client = _client(self.path)
        client.ensure_schema = lambda: None  # sheet already present

        result = client.sync_companies(
            [
                Company(
                    name="エーザイ",
                    list_url="https://www.eisai.co.jp/news/index.html",
                    source_type="html_css",
                    stock_code="4523",
                    crawl_mode="shadow",
                )
            ]
        )

        self.assertGreaterEqual(result["updated_fields"], 3)
        companies = client.load_companies()
        self.assertEqual(len(companies), 1)
        updated = companies[0]
        self.assertEqual(updated.list_url, "https://www.eisai.co.jp/news/index.html")
        self.assertEqual(updated.source_type, "html_css")
        self.assertEqual(updated.crawl_mode, "shadow")

    def test_sort_releases_by_date_sorts_data_rows_descending(self) -> None:
        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("releases")
        ws.append(RELEASE_HEADERS)
        for published_on in ("2026-07-16", "2026-07-18", "2026-07-17"):
            ws.append([published_on, "", "", "", "", "", "", "", ""])
        wb.save(self.path)

        client = _client(self.path)
        client.sort_releases_by_date()

        from openpyxl import load_workbook

        result_ws = load_workbook(self.path)["releases"]
        dates = [row[0] for row in result_ws.iter_rows(min_row=2, values_only=True)]
        self.assertEqual(dates, ["2026-07-18", "2026-07-17", "2026-07-16"])

    def test_upsert_releases_updates_existing_row_by_url(self) -> None:
        from src.models import CuratedRelease

        client = _client(self.path)
        client.ensure_schema()
        first = CuratedRelease(
            published_on=date(2026, 7, 17),
            company="協和キリン",
            title="第III相試験を開始",
            paragraph="旧リード文",
            url="https://example.com/release",
            keep=True,
        )
        client.upsert_releases([first])

        updated = first.model_copy(update={"paragraph": "新しいリード文"})
        written = client.upsert_releases([updated])

        self.assertEqual(written, 1)
        from openpyxl import load_workbook

        ws = load_workbook(self.path)["releases"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0][3], "新しいリード文")


if __name__ == "__main__":
    unittest.main()
